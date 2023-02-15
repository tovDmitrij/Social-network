###########################################################
#
# Data was taken from https://github.com/x88/i18nGeoNamesDB
#
###########################################################
import openpyxl
import db_info
import enums


def GetXLSXData(path: str, type: enums.DataType):
    '''
    Get information from Excel files

    path - path to the excel file

    type - type of data info (e.g. countries, cities, ...)
    '''
    reader = openpyxl.load_workbook(path)
    data = reader.active
    tmp = []
    sql = ""
    index = 1

    for row in range(0, data.max_row):
        print(f">>Iteration: {index} / {data.max_row}")
        index += 1
        for col in data.iter_cols(1, data.max_column):
            tmp.append(col[row].value)
        match type:
            case enums.DataType.COUNTRIES:
                sql = rf"insert into countries(id, name) values({tmp[0]}, '{tmp[1]}')"
            case enums.DataType.REGIONS:
                sql = rf"insert into regions(id, country_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
            case enums.DataType.CITIES:
                sql = rf"insert into cities(id, region_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
        try:
            db_info.DATABASE_CURSOR.execute(sql)
            db_info.DATABASE_CONNECTION.commit()
        except Exception as e:
            print(f">>Error: {e}")
        tmp.clear()