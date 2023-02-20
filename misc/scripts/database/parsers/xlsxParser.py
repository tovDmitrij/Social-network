######################################################
#
# Data of cities, regions & countries 
# was taken from https://github.com/x88/i18nGeoNamesDB
#
######################################################
import openpyxl
import db_info
import enums


def GetXLSXData(path: str, type: enums.DataType):
    '''
    Get information from Excel files and insert it into database

    path - path to the excel file

    type - type of data info (e.g. countries, cities, ...)
    '''
    READER = openpyxl.load_workbook(path)
    DATA = READER.active
    DATA_ROW_COUNT = DATA.max_row
    DATA_COLUMN_COUNT = DATA.max_column
    tmp = []
    sql = ""
    index = 1

    for row in range(0, DATA_ROW_COUNT):
        print(f">>Iteration: {index} / {DATA_ROW_COUNT}")
        index += 1

        for col in DATA.iter_cols(1, DATA_COLUMN_COUNT):
            tmp.append(col[row].value)
        match type:
            case enums.DataType.COUNTRIES:
                sql = rf"insert into countries(id, name) values({tmp[0]}, '{tmp[1]}')"
            case enums.DataType.REGIONS:
                sql = rf"insert into regions(id, country_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
            case enums.DataType.CITIES:
                sql = rf"insert into cities(id, region_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
                
        try:
            db_info.DATABASE_CURSOR.execute(sql)
            db_info.DATABASE_CONNECTION.commit()
        except Exception as e:
            print(f">>Error: {e}")
        tmp.clear()