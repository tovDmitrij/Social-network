######################################################
#
# Data of cities, regions & countries 
# was taken from https://github.com/x88/i18nGeoNamesDB
#
######################################################
import openpyxl
import db_info
import enums


def GetXLSXData(path: str, type: enums.DataType):
    '''
    Получение информации из Excel файла и её дальнейшее добавление в базу данных

    path - путь до Excel файла

    type - тип информации
    '''
    READER = openpyxl.load_workbook(path)
    DATA = READER.active
    DATA_ROW_COUNT = DATA.max_row
    DATA_COLUMN_COUNT = DATA.max_column
    tmp = []
    sql = ""

    for row in range(0, DATA_ROW_COUNT):
        for col in DATA.iter_cols(1, DATA_COLUMN_COUNT):
            tmp.append(col[row].value)
        match type:
            case enums.DataType.COUNTRIES:
                sql = rf"insert into countries(id, name) values({tmp[0]}, '{tmp[1]}')"
            case enums.DataType.REGIONS:
                sql = rf"insert into regions(id, country_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
            case enums.DataType.CITIES:
                sql = rf"insert into cities(id, region_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
        print(sql)    
        
        try:
            db_info.DATABASE_CURSOR.execute(sql)
            db_info.DATABASE_CONNECTION.commit()
        except Exception as e:
            print(f"\033[91m>>>{e}\033[0m")
        tmp.clear()