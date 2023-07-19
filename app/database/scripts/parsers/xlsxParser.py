# https://github.com/x88/i18nGeoNamesDB
import openpyxl
from db import ExecuteScript
import enums


def GetXLSXData(path: str, type: enums.Places, db: enums.Databases):
    '''
    Получение информации из Excel файла и её дальнейшее добавление в базу данных

    path - путь до Excel файла

    type - тип информации
    
    db - в какую БД вставляется информация
    '''
    READER = openpyxl.load_workbook(path)
    DATA = READER.active
    DATA_ROW_COUNT = DATA.max_row
    DATA_COLUMN_COUNT = DATA.max_column
    index = 1
    tmp = []
    sql = ""

    for row in range(0, DATA_ROW_COUNT):
        for col in DATA.iter_cols(1, DATA_COLUMN_COUNT):
            tmp.append(col[row].value)
        match type:
            case enums.Places.Countries:
                sql = rf"insert into countries(id, name) values({tmp[0]}, '{tmp[1]}')"
            case enums.Places.Regions:
                sql = rf"insert into regions(id, country_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
            case enums.Places.Cities:
                sql = rf"insert into cities(id, region_id, name) values({tmp[1]}, {tmp[0]}, '{tmp[2]}')"
        print(rf"{index} / {DATA_ROW_COUNT} - {sql}")    
        
        try:
            ExecuteScript(sql, db.value)
        except Exception as e:
            print(f"\033[91m>>>{e}\033[0m")
            return
        tmp.clear()
        index += 1